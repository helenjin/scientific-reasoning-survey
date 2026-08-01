"""One-off conversion of the working xlsx into per-sheet CSVs under data/.

Run from repo root: python3 scripts/xlsx_to_csv.py
"""
import csv
import openpyxl

SRC = "data/scientific_reasoning_resources_expanded.xlsx"

SHEET_TO_CSV = {
    "Expanded Inventory": "data/inventory.csv",
    "Coding Dimensions": "data/coding_dimensions.csv",
    "R0-R5 Taxonomy": "data/reasoning_representation_levels.csv",
    "Reasoning Error Guide": "data/reasoning_error_guide.csv",
    "New Resources Added": "data/new_resources_added.csv",
}

# Task Taxonomy sheet has two stacked tables (task definitions, then a
# "Source" grounding-citation table) sharing one sheet — split by header row.
TASK_TAXONOMY_SHEET = "Task Taxonomy"
TASK_TAXONOMY_CSV = "data/task_taxonomy.csv"
TASK_TAXONOMY_SOURCES_CSV = "data/task_taxonomy_sources.csv"

# R0-R5 sheet has a floating note in columns F-J of the header row; only the
# first 5 columns are the real table.
R0R5_MAX_COL = 5


def trim_trailing_empty(row):
    row = list(row)
    while row and row[-1] is None:
        row.pop()
    return row


def dump(ws, path, max_col=None, start_row=1, end_row=None):
    rows = list(ws.iter_rows(min_row=start_row, max_row=end_row, values_only=True))
    width = max(
        (len(trim_trailing_empty(row)) for row in rows if any(c is not None for c in row)),
        default=0,
    )
    written = 0
    with open(path, "w", newline="") as f:
        w = csv.writer(f, lineterminator="\n")
        for row in rows:
            if max_col:
                row = row[:max_col]
            if all(c is None for c in row):
                continue
            row = row[:width]
            w.writerow(["" if c is None else c for c in row])
            written += 1
    print(f"wrote {path} ({written} rows)")


def main():
    wb = openpyxl.load_workbook(SRC, data_only=True)

    for sheet, path in SHEET_TO_CSV.items():
        ws = wb[sheet]
        max_col = R0R5_MAX_COL if sheet == "R0-R5 Taxonomy" else None
        dump(ws, path, max_col=max_col)

    ws = wb[TASK_TAXONOMY_SHEET]
    rows = list(ws.iter_rows(values_only=True))
    split_at = next(i for i, row in enumerate(rows) if row[0] == "Source")
    dump(ws, TASK_TAXONOMY_CSV, end_row=split_at)
    dump(ws, TASK_TAXONOMY_SOURCES_CSV, start_row=split_at + 1)


if __name__ == "__main__":
    main()
